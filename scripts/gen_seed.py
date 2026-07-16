import openpyxl, json, random, datetime, os, re
random.seed(42)
OUT='/sessions/focused-sweet-clarke/mnt/outputs/avanco-obras/public'
wb_m = openpyxl.load_workbook('Avanço Fisico - M01.xlsx', data_only=True, read_only=True)
ws = wb_m['RESUMO DETALHADO']
services=[]
for row in ws.iter_rows(min_row=3, max_row=70, values_only=True):
    if row[1] and row[2] and str(row[1]).strip().isdigit():
        services.append({'item':int(row[1]),'nome':str(row[2]).strip().title(),'equipe':str(row[4]).strip() if row[4] else 'Geral'})
seen=set(); svc=[]
for s in services:
    if s['nome'] in seen: continue
    seen.add(s['nome']); svc.append(s)
for i,s in enumerate(svc): s['item']=i+1

apt_map={'Bl. A - 01':'A','Bl. C - 02':'C','Bl. E - 03':'E','Bl. G - 04':'G','Bl. H - 05':'H','Bl. F - 06':'F','Bl. D - 07':'D','Bl. B - 08':'B'}
blocos=[]
for sheet,bl in apt_map.items():
    ws2=wb_m[sheet]
    r3=list(ws2.iter_rows(min_row=3,max_row=3,values_only=True))[0]
    apts=[(str(c).zfill(3) if str(c).isdigit() else str(c)) for c in r3 if c not in (None,'')]
    blocos.append({'bloco':bl,'apts':apts})

wb_c=openpyxl.load_workbook('Avanço Fisico - C03.xlsx', data_only=True, read_only=True)
conj={}
for sh in ['Conj. A - B','Conj. C - D - E - F - G','Conj. H','Conj. I - J - K','Conj. L - M','Conj. N - O - P - Q - R']:
    ws3=wb_c[sh]
    for r in ws3.iter_rows(min_row=1,max_row=6,values_only=True):
        for c in r:
            if isinstance(c,str):
                m=re.fullmatch(r'([A-R])(\d{2})',c.strip())
                if m: conj.setdefault(m.group(1),set()).add(m.group(2))
conjuntos=[{'conjunto':k,'casas':sorted(v)} for k,v in sorted(conj.items())]

def status_from(p):
    return 'concluido' if p>=100 else ('execucao' if p>=1 else 'nao_iniciado')
def gen_services(base):
    out=[]
    for s in svc:
        w=base*140 - s['item']*1.1 + random.uniform(-15,15)
        p=round(min(100,max(0,w)))
        st=status_from(p)
        if 0<p<100 and random.random()<0.06: st='atencao'
        if 0<p<100 and random.random()<0.04: st='atrasado'
        out.append({'servicoItem':s['item'],'nome':s['nome'],'equipe':s['equipe'],'percentual':p,'status':st,'responsavel':random.choice(['J. Silva','M. Souza','R. Lima','C. Alves','P. Rocha']),'data':'2026-07-09','obs':''})
    return out
def unit_pct(sv): return round(sum(x['percentual'] for x in sv)/len(sv),1) if sv else 0

units=[]
for b in blocos:
    bias=random.uniform(0.35,0.85)
    for apt in b['apts']:
        sv=gen_services(bias+random.uniform(-0.1,0.1)); pct=unit_pct(sv)
        units.append({'id':f'M01-{b["bloco"]}-{apt}','condominio':'M01','condominioNome':'Alto do Jerivá','tipo':'apartamento','bloco':b['bloco'],'conjunto':None,'unidade':apt,'percentual':pct,'status':status_from(pct),'servicos':sv})
for c in conjuntos:
    bias=random.uniform(0.3,0.9)
    for casa in c['casas']:
        sv=gen_services(bias+random.uniform(-0.12,0.12)); pct=unit_pct(sv)
        units.append({'id':f'C03-{c["conjunto"]}-{casa}','condominio':'C03','condominioNome':'Alto do Burití','tipo':'casa','bloco':None,'conjunto':c['conjunto'],'unidade':casa,'percentual':pct,'status':status_from(pct),'servicos':sv})

weeks=['2026-06-11','2026-06-18','2026-06-25','2026-07-02','2026-07-09']
def cond_avg(cid,f):
    us=[u for u in units if u['condominio']==cid]; return round(sum(u['percentual'] for u in us)/len(us)*f,1)
historico=[]
for i,wk in enumerate(weeks):
    f=0.72+i*0.06
    for cid in ['C03','M01']:
        historico.append({'semana':wk,'condominio':cid,'percentualMedio':cond_avg(cid,f)})

data={'geradoEm':datetime.datetime.now().isoformat(),
 'condominios':[
   {'id':'C03','nome':'Alto do Burití','tipo':'casa','totalUnidades':len([u for u in units if u['condominio']=='C03'])},
   {'id':'M01','nome':'Alto do Jerivá','tipo':'apartamento','totalUnidades':len([u for u in units if u['condominio']=='M01'])}],
 'servicos':svc,'conjuntos':conjuntos,
 'blocos':[{'bloco':b['bloco'],'totalApts':len(b['apts'])} for b in blocos],
 'unidades':units,'historico':historico}
json.dump(data,open(OUT+'/seed.json','w'),ensure_ascii=False)
print('services:',len(svc),'conjuntos:',len(conjuntos),'blocos:',len(blocos),'unidades:',len(units))
print('C03:',len([u for u in units if u["condominio"]=="C03"]),'M01:',len([u for u in units if u["condominio"]=="M01"]))
print('KB:',round(os.path.getsize(OUT+'/seed.json')/1024))
